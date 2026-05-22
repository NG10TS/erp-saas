"""
Generador de Reportes Excel (.xlsx)
Usa openpyxl para crear archivos Excel profesionales
"""
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import logging

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Genera reportes en formato Excel con diseño profesional"""
    
    # ─── Estilos ────────────────────────────────────────────
    HEADER_FONT = Font(name='Inter', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    TITLE_FONT = Font(name='Inter', size=16, bold=True, color='1B5E20')
    SUBTITLE_FONT = Font(name='Inter', size=10, color='666666')
    
    DATA_FONT = Font(name='Inter', size=10)
    DATA_ALIGNMENT = Alignment(vertical='center')
    
    MONEY_FORMAT = '#,##0.00'
    DATE_FORMAT = 'DD/MM/YYYY'
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0'),
    )
    
    ALTERNATE_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    TOTAL_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    TOTAL_FONT = Font(name='Inter', size=11, bold=True)
    
    def __init__(self, company_name: str = "ERP Conversacional"):
        self.company_name = company_name
    
    def generate_sales_report(
        self,
        data: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Genera reporte de ventas en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # ─── Título ─────────────────────────────────────────
        self._add_title(ws, "Reporte de Ventas", filters)
        
        # ─── Headers ────────────────────────────────────────
        headers = [
            'N° Venta', 'Fecha', 'Cliente', 'Estado', 'Pago',
            'Método', 'Subtotal', 'Descuento', 'IVA', 'Total'
        ]
        self._add_headers(ws, headers, row=4)
        
        # ─── Datos ──────────────────────────────────────────
        for i, row in enumerate(data, start=5):
            values = [
                row.get('numero_venta', ''),
                row.get('fecha_venta', ''),
                row.get('customer_name', ''),
                row.get('estado', ''),
                row.get('estado_pago', ''),
                row.get('metodo_pago', ''),
                row.get('subtotal', 0),
                row.get('descuento', 0),
                row.get('iva', 0),
                row.get('total', 0),
            ]
            self._add_data_row(ws, i, values)
        
        # ─── Totales ────────────────────────────────────────
        last_row = 4 + len(data) + 1
        self._add_totals_row(ws, last_row, headers, data)
        
        # ─── Auto-ajustar columnas ──────────────────────────
        self._auto_fit_columns(ws)
        
        # ─── Guardar ────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def generate_inventory_report(
        self,
        data: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Genera reporte de inventario en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        self._add_title(ws, "Reporte de Inventario", filters)
        
        headers = [
            'SKU', 'Producto', 'Categoría', 'Stock Actual',
            'Stock Mínimo', 'Precio Venta', 'Costo', 'Utilidad', 'Estado'
        ]
        self._add_headers(ws, headers, row=4)
        
        for i, row in enumerate(data, start=5):
            values = [
                row.get('sku', ''),
                row.get('name', ''),
                row.get('category_name', ''),
                row.get('stock_actual', 0),
                row.get('stock_minimo', 0),
                row.get('precio_venta', 0),
                row.get('costo', 0),
                row.get('utilidad', 0),
                'Activo' if row.get('is_active') else 'Inactivo',
            ]
            self._add_data_row(ws, i, values)
        
        self._auto_fit_columns(ws)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def generate_customers_report(
        self,
        data: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Genera reporte de clientes en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        self._add_title(ws, "Reporte de Clientes", filters)
        
        headers = [
            'Nombre', 'Identificación', 'Email', 'Teléfono',
            'Total Compras', 'Total Gastado', 'Última Compra'
        ]
        self._add_headers(ws, headers, row=4)
        
        for i, row in enumerate(data, start=5):
            values = [
                row.get('name', ''),
                row.get('identification', ''),
                row.get('email', ''),
                row.get('phone_number', ''),
                row.get('total_purchases', 0),
                row.get('total_spent', 0),
                row.get('last_purchase', ''),
            ]
            self._add_data_row(ws, i, values)
        
        self._auto_fit_columns(ws)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def generate_invoices_report(
        self,
        data: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Genera reporte de facturas en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas SRI"
        
        self._add_title(ws, "Reporte de Facturas Electrónicas", filters)
        
        headers = [
            'N° Factura', 'Secuencial', 'Fecha', 'Cliente',
            'Identificación', 'Subtotal', 'IVA', 'Total',
            'Estado SRI', 'Autorización', 'Fecha Autorización'
        ]
        self._add_headers(ws, headers, row=4)
        
        for i, row in enumerate(data, start=5):
            values = [
                row.get('invoice_number', '')[:20] if row.get('invoice_number') else '',
                row.get('sequential', ''),
                row.get('issue_date', ''),
                row.get('customer_name', ''),
                row.get('customer_identification', ''),
                row.get('subtotal', 0),
                row.get('iva', 0),
                row.get('total', 0),
                row.get('sri_status', ''),
                row.get('authorization_number', ''),
                row.get('authorization_date', ''),
            ]
            self._add_data_row(ws, i, values)
        
        self._auto_fit_columns(ws)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def generate_iva_report(
        self,
        data: List[Dict[str, Any]],
        year: int,
        filters: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Genera reporte de IVA (formato SRI) en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"IVA {year}"
        
        self._add_title(ws, f"Resumen de IVA - {year}", filters)
        
        headers = [
            'Año', 'Mes', 'Total Ventas', 'Base Imponible',
            'IVA Cobrado', 'Total Facturado'
        ]
        self._add_headers(ws, headers, row=4)
        
        for i, row in enumerate(data, start=5):
            values = [
                row.get('year', ''),
                row.get('month', ''),
                row.get('total_sales', 0),
                row.get('base_imponible', 0),
                row.get('iva_cobrado', 0),
                row.get('total_facturado', 0),
            ]
            self._add_data_row(ws, i, values)
        
        self._auto_fit_columns(ws)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # ─── Métodos privados ──────────────────────────────────
    
    def _add_title(self, ws, title: str, filters: Optional[Dict] = None):
        """Agrega título y metadatos al reporte"""
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 35
        
        # Empresa y fecha
        ws.merge_cells('A2:J2')
        ws['A2'].value = f"{self.company_name} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = self.SUBTITLE_FONT
        
        # Filtros aplicados
        if filters:
            ws.merge_cells('A3:J3')
            filter_text = "Filtros: " + " | ".join(
                f"{k}: {v}" for k, v in filters.items() if v
            )
            ws['A3'].value = filter_text
            ws['A3'].font = Font(name='Inter', size=9, color='999999')
    
    def _add_headers(self, ws, headers: List[str], row: int = 4):
        """Agrega fila de encabezados con estilo"""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        ws.row_dimensions[row].height = 30
    
    def _add_data_row(self, ws, row: int, values: List[Any]):
        """Agrega fila de datos con formato"""
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.DATA_FONT
            cell.alignment = self.DATA_ALIGNMENT
            cell.border = self.THIN_BORDER
            
            # Formato moneda para columnas numéricas
            if isinstance(value, (int, float)) and col >= 7:
                cell.number_format = self.MONEY_FORMAT
            
            # Colores alternados
            if row % 2 == 0:
                cell.fill = self.ALTERNATE_FILL
    
    def _add_totals_row(self, ws, row: int, headers: List[str], data: List[Dict]):
        """Agrega fila de totales"""
        total_subtotal = sum(d.get('subtotal', 0) for d in data)
        total_iva = sum(d.get('iva', 0) for d in data)
        total_general = sum(d.get('total', 0) for d in data)
        
        totals = ['', '', '', '', '', 'TOTALES:', total_subtotal, '', total_iva, total_general]
        
        for col, value in enumerate(totals, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = self.TOTAL_FONT
            cell.fill = self.TOTAL_FILL
            cell.border = self.THIN_BORDER
            if isinstance(value, (int, float)):
                cell.number_format = self.MONEY_FORMAT
    
    def _auto_fit_columns(self, ws):
        """Auto-ajusta el ancho de columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 40)
            ws.column_dimensions[column_letter].width = adjusted_width