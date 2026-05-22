"""
Report Service
Generates Excel and CSV exports for sales and inventory.
Uses openpyxl for Excel (already in requirements.txt).
"""
from datetime import date, datetime, timedelta
from typing import Optional
from uuid import UUID
from io import BytesIO
import logging

from sqlalchemy.orm import Session
from sqlalchemy import and_, cast, Date, func

from app.models.sale import Sale, SaleItem, SaleStatus
from app.models.product import Product
from app.models.customer import Customer
from app.models.invoice import Invoice

logger = logging.getLogger(__name__)


class ReportService:
    """Generate downloadable reports for the business"""

    def __init__(self, db: Session, business_id: UUID):
        self.db = db
        self.business_id = business_id

    # ─────────────────────────────────────────────────────────────────────────
    # Sales Excel report
    # ─────────────────────────────────────────────────────────────────────────

    def export_sales_excel(
        self,
        start_date: Optional[date] = None,
        end_date:   Optional[date] = None,
    ) -> bytes:
        """
        Generate an Excel file with completed sales in the date range.

        Args:
            start_date: Filter from (default: start of current month)
            end_date:   Filter to (default: today)

        Returns:
            Raw .xlsx bytes suitable for a FileResponse
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not start_date:
            start_date = date.today().replace(day=1)
        if not end_date:
            end_date = date.today()

        # ── Query ────────────────────────────────────────────────────────────
        rows = (
            self.db.query(
                Sale.numero_venta,
                Sale.fecha_venta,
                Sale.total,
                Sale.subtotal,
                Sale.iva,
                Sale.descuento,
                Sale.metodo_pago,
                Sale.estado,
                Customer.name.label("customer_name"),
                Customer.identification.label("customer_id"),
            )
            .outerjoin(Customer, Sale.customer_id == Customer.id)
            .filter(
                and_(
                    Sale.business_id == self.business_id,
                    cast(Sale.fecha_venta, Date) >= start_date,
                    cast(Sale.fecha_venta, Date) <= end_date,
                )
            )
            .order_by(Sale.fecha_venta.asc())
            .all()
        )

        # ── Workbook setup ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"

        # Colors
        HEADER_FILL  = PatternFill("solid", fgColor="075E54")   # WhatsApp green
        ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")
        TOTAL_FILL   = PatternFill("solid", fgColor="DCF8C6")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Title ─────────────────────────────────────────────────────────────
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = f"Reporte de Ventas — {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="064E47")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Headers ───────────────────────────────────────────────────────────
        headers = [
            "N° Venta", "Fecha", "Cliente", "RUC/Cédula",
            "Subtotal", "Descuento", "IVA", "Total",
            "Método Pago", "Estado",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font       = Font(bold=True, color="FFFFFF", size=10)
            cell.fill       = HEADER_FILL
            cell.alignment  = Alignment(horizontal="center", vertical="center")
            cell.border     = border
        ws.row_dimensions[2].height = 20

        # ── Data rows ─────────────────────────────────────────────────────────
        payment_labels = {
            "cash": "Efectivo", "card": "Tarjeta",
            "transfer": "Transferencia", "qr": "QR",
        }
        status_labels = {
            "completed": "Completada", "cancelled": "Cancelada",
            "pending": "Pendiente", "confirmed": "Confirmada",
        }

        for row_idx, sale in enumerate(rows, 3):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            values = [
                sale.numero_venta,
                sale.fecha_venta.strftime("%d/%m/%Y %H:%M") if sale.fecha_venta else "",
                sale.customer_name or "Consumidor Final",
                sale.customer_id or "9999999999999",
                float(sale.subtotal or 0),
                float(sale.descuento or 0),
                float(sale.iva or 0),
                float(sale.total or 0),
                payment_labels.get(str(sale.metodo_pago or "cash").lower(), str(sale.metodo_pago or "")),
                status_labels.get(str(sale.estado or "").lower(), str(sale.estado or "")),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if isinstance(value, float):
                    cell.number_format = '"$"#,##0.00'

        # ── Totals row ────────────────────────────────────────────────────────
        total_row = len(rows) + 3
        ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = TOTAL_FILL

        for col, field in [(5, "subtotal"), (6, "descuento"), (7, "iva"), (8, "total")]:
            total = sum(float(getattr(r, field) or 0) for r in rows)
            cell = ws.cell(row=total_row, column=col, value=total)
            cell.font         = Font(bold=True)
            cell.fill         = TOTAL_FILL
            cell.number_format = '"$"#,##0.00'
            cell.border       = border

        # ── Column widths ─────────────────────────────────────────────────────
        widths = [15, 18, 25, 15, 12, 12, 12, 12, 14, 12]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Resumen")
        ws2["A1"] = "Período"
        ws2["B1"] = f"{start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')}"
        ws2["A2"] = "Total ventas"
        ws2["B2"] = len(rows)
        ws2["A3"] = "Total ingresos"
        ws2["B3"] = sum(float(r.total or 0) for r in rows)
        ws2["B3"].number_format = '"$"#,##0.00'
        ws2["A4"] = "Generado"
        ws2["B4"] = datetime.now().strftime("%d/%m/%Y %H:%M")

        for row in ws2.iter_rows():
            for cell in row:
                cell.font = Font(size=11)

        # ── Export ────────────────────────────────────────────────────────────
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    # ─────────────────────────────────────────────────────────────────────────
    # Inventory CSV
    # ─────────────────────────────────────────────────────────────────────────

    def export_inventory_csv(self) -> bytes:
        """
        Generate a CSV with current inventory status.

        Returns:
            UTF-8 bytes (CSV with BOM for Excel compatibility)
        """
        import csv
        from io import StringIO

        products = (
            self.db.query(
                Product.sku,
                Product.name,
                Product.stock_actual,
                Product.stock_minimo,
                Product.precio_venta,
            )
            .filter(
                and_(
                    Product.business_id == self.business_id,
                    Product.is_active == True,
                )
            )
            .order_by(Product.name.asc())
            .all()
        )

        output = StringIO()
        writer = csv.writer(output)

        writer.writerow(["SKU", "Producto", "Stock Actual", "Stock Mínimo", "Precio Venta", "Estado"])

        for p in products:
            stock_status = (
                "CRÍTICO" if p.stock_actual == 0
                else "BAJO" if p.stock_actual <= p.stock_minimo
                else "OK"
            )
            writer.writerow([
                p.sku or "",
                p.name,
                p.stock_actual,
                p.stock_minimo,
                f"{float(p.precio_venta):.2f}",
                stock_status,
            ])

        # BOM for Excel UTF-8 detection
        return ("\ufeff" + output.getvalue()).encode("utf-8")